import streamlit as st
import pandas as pd
import yaml
import os
from pathlib import Path
import plotly.graph_objects as go
import numpy as np

from src.parser import parse_qvm_content, parse_filename, QVMParseError
from src.calculations import calculate_annular_ring, calculate_cam_compensation
from src.visuals import plot_bullseye_scatter, plot_quiver, plot_heatmap
from panel_mapping import create_four_quarters_view

# Enable Pandas Copy-on-Write for performance
pd.options.mode.copy_on_write = True

def load_settings():
    config_path = os.path.join(os.path.dirname(__file__), 'config', 'settings.yaml')
    with open(config_path, 'r') as f:
        return yaml.safe_load(f)

def load_css(file_path: str, settings: dict = None) -> None:
    """Loads a CSS file and injects it into the Streamlit app."""
    try:
        css = Path(file_path).read_text()
        colors = (settings or {}).get('COLORS', {})
        bg      = colors.get('app_background', '#212121')
        text    = colors.get('app_text', '#FFFFFF')
        panel   = colors.get('panel_ui', '#B87333')
        hover   = colors.get('panel_ui_hover', '#d48c46')
        css_variables = f'''
        <style>
            :root {{
                --background-color: {bg};
                --text-color: {text};
                --panel-color: {panel};
                --panel-hover-color: {hover};
            }}
            {css}
        </style>
        '''
        st.markdown(css_variables, unsafe_allow_html=True)
    except FileNotFoundError:
        pass

def highlight_annular_ring(row, col_name, threshold):
    """Pandas styler function to highlight rows with low annular ring."""
    if pd.isna(row[col_name]):
        return [''] * len(row)
    if row[col_name] < threshold:
        return ['background-color: #ffcccc; color: #990000; font-weight: bold'] * len(row)
    return [''] * len(row)

def _nav_buttons(options, state_key, default=None, cols_gap="small"):
    """Render a row of primary/secondary buttons for navigation."""
    if state_key not in st.session_state or st.session_state[state_key] not in options:
        st.session_state[state_key] = default if default is not None else options[0]
    cols = st.columns(len(options), gap=cols_gap)
    for i, label in enumerate(options):
        is_active = st.session_state[state_key] == label
        cols[i].button(
            str(label),
            type="primary" if is_active else "secondary",
            width="stretch",
            key=f"{state_key}_{label}",
            on_click=lambda l=label: st.session_state.update({state_key: l}),
        )
    return st.session_state[state_key]


def main():
    settings = load_settings()
    ui_strings = settings.get('UI_STRINGS', {})
    col_names = settings.get('COLUMN_NAMES', {})
    tolerances = settings.get('TOLERANCES', {})
    plot_settings = settings.get('PLOT_SETTINGS', {})
    chart_colors = settings.get('CHART_COLORS', {})
    chart_heights = settings.get('CHART_HEIGHTS', {})
    optical_thresholds = settings.get('OPTICAL_EDGE_THRESHOLDS', {})
    chart_markers = settings.get('CHART_MARKERS', {})

    st.set_page_config(page_title=ui_strings.get('dashboard_title', 'QVM Dashboard'), layout="wide")
    load_css(os.path.join(os.path.dirname(__file__), "assets", "styles.css"), settings)

    st.title(ui_strings.get('dashboard_title', 'QVM Dashboard'))

    # --- Sidebar ---
    with st.sidebar:
        st.header("File Upload & Meta Data")
        uploaded_files = st.file_uploader(
            "Upload QVM Text Files",
            type=['txt'],
            accept_multiple_files=True
        )

        if uploaded_files:
            if st.button("🔄 Run Analysis", type="primary", use_container_width=True):
                st.session_state['run_analysis'] = True

        st.markdown("---")
        material_build_up = st.text_input("Material Build-up", "")
        batch_id = st.text_input("Batch ID", "")
        
        # --- Process Stability Limits Configuration ---
        st.markdown("---")
        st.subheader("📊 Process Limits (µm)")
        
        with st.expander("🔴 Via Drill Limits", expanded=False):
            default_via_nom = settings.get('PROCESS_STABILITY', {}).get('via_nominal_diameter', 0.020) * 1000
            default_via_ucl = settings.get('PROCESS_STABILITY', {}).get('via_ucl', 0.0215) * 1000
            default_via_lcl = settings.get('PROCESS_STABILITY', {}).get('via_lcl', 0.0185) * 1000
            
            via_nom = st.number_input("Via Nominal (µm)", value=default_via_nom, step=0.1, key="via_nom_input")
            via_ucl = st.number_input("Via UCL (µm)", value=default_via_ucl, step=0.1, key="via_ucl_input")
            via_lcl = st.number_input("Via LCL (µm)", value=default_via_lcl, step=0.1, key="via_lcl_input")
            
            # Store in session state
            st.session_state['via_nominal'] = via_nom / 1000
            st.session_state['via_ucl'] = via_ucl / 1000
            st.session_state['via_lcl'] = via_lcl / 1000
        
        with st.expander("🟠 Pad Etch Limits", expanded=False):
            default_pad_nom = settings.get('PROCESS_STABILITY', {}).get('pad_nominal_diameter', 0.3125) * 1000
            default_pad_ucl = settings.get('PROCESS_STABILITY', {}).get('pad_ucl', 0.3157) * 1000
            default_pad_lcl = settings.get('PROCESS_STABILITY', {}).get('pad_lcl', 0.3093) * 1000
            
            pad_nom = st.number_input("Pad Nominal (µm)", value=default_pad_nom, step=0.1, key="pad_nom_input")
            pad_ucl = st.number_input("Pad UCL (µm)", value=default_pad_ucl, step=0.1, key="pad_ucl_input")
            pad_lcl = st.number_input("Pad LCL (µm)", value=default_pad_lcl, step=0.1, key="pad_lcl_input")
            
            # Store in session state
            st.session_state['pad_nominal'] = pad_nom / 1000
            st.session_state['pad_ucl'] = pad_ucl / 1000
            st.session_state['pad_lcl'] = pad_lcl / 1000

    # --- Data Ingestion ---
    all_data = []
    available_panels = set()
    available_sides = set()

    if uploaded_files:
        for file in uploaded_files:
            try:
                # Extract meta from filename
                filename_info = parse_filename(file.name)
                panel_id = filename_info.get('Panel Number', 'Unknown')
                side_id = filename_info.get('Side', 'Unknown')

                content = file.getvalue().decode('utf-8')
                df = parse_qvm_content(content)
                df = calculate_annular_ring(df, settings)

                # Tag dataframe with meta
                df['Panel'] = panel_id
                df['Side'] = side_id

                available_panels.add(panel_id)
                available_sides.add(side_id)
                all_data.append(df)
            except Exception as e:
                st.sidebar.error(f"Error parsing {file.name}: {e}")

    # --- Main UI ---
    # Top Level Navigation
    main_view = _nav_buttons(
        ["Panel Map", "Pad to Via", "Via to Pad", "Alignment"],
        state_key="main_view",
        default="Panel Map",
    )

    # Panel Map requires no uploaded data
    if main_view == "Panel Map":
        fig = create_four_quarters_view(settings)
        st.plotly_chart(fig, use_container_width=True)
        return

    if not all_data:
        st.info("Please upload one or more QVM text log files in the sidebar to begin.")
        return

    # Combine all parsed data
    master_df = pd.concat(all_data, ignore_index=True)

    if main_view == "Pad to Via":
        # Expander for Scope
        with st.expander("Analysis Scope", expanded=True):
            col1, col2 = st.columns(2)
            with col1:
                panel_list = sorted(list(available_panels))
                if len(panel_list) > 1:
                    panel_list.insert(0, "All")
                # Create display names with "Panel-" prefix
                panel_display = ["All" if p == "All" else f"Panel-{p}" for p in panel_list]
                st.markdown("**Select Panel**")
                selected_panel_display = _nav_buttons(panel_display, state_key="selected_panel", default=panel_display[0])
                # Map display back to actual panel ID
                selected_panel = "All" if selected_panel_display == "All" else selected_panel_display.replace("Panel-", "")

            with col2:
                side_list = sorted(list(available_sides))
                if len(side_list) > 1:
                    side_list.insert(0, "Both")
                st.markdown("**Select Side**")
                selected_side = _nav_buttons(side_list, state_key="selected_side", default=side_list[0])

        # Filter Data
        filtered_df = master_df.copy()
        if selected_panel and selected_panel != "All":
            filtered_df = filtered_df[filtered_df['Panel'] == selected_panel]
        if selected_side and selected_side != "Both":
            filtered_df = filtered_df[filtered_df['Side'] == selected_side]

        if filtered_df.empty:
            st.warning("No data matches the selected filters.")
            return

        st.markdown("<br>", unsafe_allow_html=True)

        # Sub-level navigation
        sub_view = _nav_buttons(
            ["Quality Control", "Analytics", "CAM Compensation", "Optical Edge Confidence", "Process Stability"],
            state_key="sub_view",
            default="Quality Control",
        )

        st.markdown("---")

        # --- Quality Control View ---
        if sub_view == "Quality Control":
            annular_col = col_names.get('annular_ring', 'Annular Ring')
            threshold = tolerances.get('annular_ring_min', 0.0)
            
            # Prepare display dataframe
            display_df = filtered_df.copy()
            
            # Drop Panel and Side columns (not needed in display)
            cols_to_drop = [col for col in ['Panel', 'Side'] if col in display_df.columns]
            display_df = display_df.drop(columns=cols_to_drop)
            
            # Convert all measurements from mm to microns (multiply by 1000)
            measurement_cols = [
                col_names.get('outer_diameter', 'Outer Diameter'),
                col_names.get('inner_diameter', 'Inner Diameter'),
                col_names.get('ptv_distance', 'PtV Distance'),
                col_names.get('x_distance', 'Shift (DX)'),
                col_names.get('y_distance', 'Shift (DY)'),
                col_names.get('annular_ring', 'Annular Ring')
            ]
            for col in measurement_cols:
                if col in display_df.columns:
                    display_df[col] = (display_df[col] * 1000).round(3)
            
            # Sort by Grid ID ascending
            grid_id_col = col_names.get('grid_id', 'Grid ID')
            if grid_id_col in display_df.columns:
                display_df = display_df.sort_values(by=grid_id_col, ascending=True)
            
            # Statistical Summary Panel
            st.subheader("📊 Statistical Summary")
            summary_cols = st.columns(5)
            
            # Calculate stats for annular ring
            if annular_col in display_df.columns:
                ar_min = display_df[annular_col].min()
                ar_max = display_df[annular_col].max()
                ar_mean = display_df[annular_col].mean()
                ar_std = display_df[annular_col].std()
                fail_count = (display_df[annular_col] < threshold).sum()
                fail_pct = (fail_count / len(display_df) * 100) if len(display_df) > 0 else 0
                
                with summary_cols[0]:
                    st.metric("Min Annular Ring", f"{ar_min:.3f} µm", delta=f"{ar_min - threshold:.3f}" if ar_min < threshold else None, delta_color="inverse")
                with summary_cols[1]:
                    st.metric("Max Annular Ring", f"{ar_max:.3f} µm")
                with summary_cols[2]:
                    st.metric("Mean ± StdDev", f"{ar_mean:.3f} ± {ar_std:.3f} µm")
                with summary_cols[3]:
                    worst_idx = display_df[annular_col].idxmin()
                    worst_grid = display_df.loc[worst_idx, grid_id_col] if grid_id_col in display_df.columns else 'N/A'
                    st.metric("Worst Point", f"Grid {worst_grid}", f"{ar_min:.3f} µm")
                with summary_cols[4]:
                    st.metric("Failures", f"{fail_count} ({fail_pct:.1f}%)", 
                             delta=f"{fail_count} points" if fail_count > 0 else "✓ All Pass",
                             delta_color="inverse" if fail_count > 0 else "off")
            
            st.markdown("---")
            
            # Export to Excel button
            col_left, col_right = st.columns([3, 1])
            with col_left:
                st.write(f"Highlighting rows where {annular_col} < {threshold}")
            with col_right:
                # Create Excel file in memory
                from io import BytesIO
                buffer = BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    # Write data with highlighting
                    display_df_export = display_df.copy()
                    display_df_export.to_excel(writer, sheet_name='QVM Data', index=False)
                    
                    # Get workbook and worksheet
                    workbook = writer.book
                    worksheet = writer.sheets['QVM Data']
                    
                    # Apply conditional formatting for failed points
                    from openpyxl.styles import PatternFill
                    red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                    
                    if annular_col in display_df_export.columns:
                        ar_col_idx = list(display_df_export.columns).index(annular_col) + 1  # Excel is 1-indexed
                        for row_idx, value in enumerate(display_df_export[annular_col], start=2):  # start=2 to skip header
                            if value < threshold:
                                for col_idx in range(1, len(display_df_export.columns) + 1):
                                    worksheet.cell(row=row_idx, column=col_idx).fill = red_fill
                
                buffer.seek(0)
                st.download_button(
                    label="📥 Export to Excel",
                    data=buffer,
                    file_name=f"QVM_Quality_Report_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary"
                )
            
            # Add clarification about Grid ID
            st.info("📍 **Grid ID**: Position on the PCB quadrant (11-14 = Upper Left, 21-24 = Lower Left, 31-34 = Lower Right, 41-44 = Upper Right). First digit = quadrant, second digit = point (1-4). All measurements in **microns (µm)**.")

            # Apply styling with custom formatter for 3 decimals without trailing zeros
            def format_to_3_decimals(val):
                if pd.isna(val):
                    return ''
                try:
                    # Format to 3 decimals, then strip trailing zeros and decimal point if needed
                    formatted = f'{float(val):.3f}'.rstrip('0').rstrip('.')
                    return formatted
                except:
                    return str(val)
            
            # Build format dict for all numeric columns at once
            format_dict = {}
            for col in display_df.columns:
                if display_df[col].dtype in ['float64', 'float32', 'int64', 'int32']:
                    format_dict[col] = format_to_3_decimals
            
            styled_df = display_df.style.apply(
                highlight_annular_ring,
                col_name=annular_col,
                threshold=threshold,
                axis=1
            )
            if format_dict:
                styled_df = styled_df.format(format_dict)
            st.dataframe(styled_df, use_container_width=True)

        # --- Analytics View ---
        elif sub_view == "Analytics":
            plot_type = st.selectbox(
                "Select Plot Type",
                ["Bullseye Scatter", "Quiver/Vector Plot", "Heatmap"]
            )

            if plot_type == "Bullseye Scatter":
                fig = plot_bullseye_scatter(filtered_df, settings)
                st.plotly_chart(fig, use_container_width=True, height=chart_heights.get('analytics_plot', 600))

            elif plot_type == "Quiver/Vector Plot":
                multiplier = st.slider(
                    "Vector Exaggeration Multiplier",
                    min_value=plot_settings.get('vector_multiplier_min', 1),
                    max_value=plot_settings.get('vector_multiplier_max', 200),
                    value=plot_settings.get('vector_multiplier_default', 50)
                )
                fig = plot_quiver(filtered_df, settings, multiplier)
                st.plotly_chart(fig, use_container_width=True, height=chart_heights.get('analytics_plot', 600))

            elif plot_type == "Heatmap":
                fig = plot_heatmap(filtered_df, settings)
                st.plotly_chart(fig, use_container_width=True, height=chart_heights.get('analytics_plot', 600))

        # --- CAM Compensation View ---
        elif sub_view == "CAM Compensation":
            avg_x, avg_y = calculate_cam_compensation(filtered_df, settings)
            col1, col2 = st.columns(2)
            with col1:
                st.metric("Average X Shift (DX)", f"{avg_x:.6f} mm")
            with col2:
                st.metric("Average Y Shift (DY)", f"{avg_y:.6f} mm")

        elif sub_view == "Optical Edge Confidence":
            st.subheader("🔍 Optical Edge Confidence Analysis")
            st.write("Analyze Point Count (Pts.) metric from optical inspection for hole integrity and plating quality assessment.")
            
            # Check if 'Pts.' column exists
            pts_col = 'Pts.'
            if pts_col not in filtered_df.columns:
                st.error(f"Column '{pts_col}' not found in data. Available columns: {list(filtered_df.columns)}")
                return
            
            # --- 1. Via Health Bar Chart ---
            st.markdown("### 1️⃣ Via Health Status (Traffic Light Logic)")
            
            # Filter for Vias only
            via_df = filtered_df[filtered_df['Type'] == 'Via'].copy() if 'Type' in filtered_df.columns else pd.DataFrame()
            
            if not via_df.empty:
                # Prepare data for bar chart with Grid ID
                grid_id_col = col_names.get('grid_id', 'Grid ID')
                via_data = via_df.groupby('Location')[pts_col].mean().reset_index()
                # Get first Grid ID value for each location
                grid_ids = via_df.groupby('Location')[grid_id_col].first().reset_index()
                via_data = via_data.merge(grid_ids, on='Location')
                via_data.columns = ['Via ID', 'Point Count', 'Grid ID']
                
                # Define colors based on traffic light logic
                def get_health_color(points):
                    if pd.isna(points):
                        return chart_colors.get('traffic_light_gray', '#999999')
                    green_threshold = optical_thresholds.get('green_min', 75)
                    orange_threshold = optical_thresholds.get('orange_min', 40)
                    if points > green_threshold:
                        return chart_colors.get('traffic_light_green', '#2ecc71')
                    elif points >= orange_threshold:
                        return chart_colors.get('traffic_light_orange', '#f39c12')
                    else:
                        return chart_colors.get('traffic_light_red', '#e74c3c')
                
                via_data['Color'] = via_data['Point Count'].apply(get_health_color)
                
                # Create bar chart with Grid ID labels
                fig_via_health = go.Figure(data=[
                    go.Bar(
                        x=via_data['Via ID'],
                        y=via_data['Point Count'],
                        marker=dict(color=via_data['Color']),
                        text=[f"{pts:.1f}<br>({gid})" for pts, gid in zip(via_data['Point Count'], via_data['Grid ID'])],
                        textposition='outside',
                        customdata=via_data['Grid ID'],
                        hovertemplate='<b>%{x}</b><br>Grid ID: %{customdata}<br>Points: %{y:.1f}<extra></extra>'
                    )
                ])
                
                fig_via_health.update_layout(
                    title="Via Health Status by Point Count",
                    xaxis_title="Via ID",
                    yaxis_title="Point Count (Pts.)",
                    height=chart_heights.get('health_status', 550),
                    showlegend=False,
                    plot_bgcolor=chart_colors.get('chart_background', '#FFFFFF'),
                    hovermode='x unified',
                    xaxis=dict(showgrid=chart_colors.get('chart_gridlines_visible', False)),
                    yaxis=dict(showgrid=chart_colors.get('chart_gridlines_visible', False))
                )
                
                st.plotly_chart(fig_via_health, use_container_width=True)
            else:
                st.warning("No Via data found in the dataset.")
            
            # --- 1a. Pad Health Bar Chart ---
            st.markdown("### 1️⃣a Pad Health Status (Traffic Light Logic)")
            
            # Filter for Pads only
            pad_df = filtered_df[filtered_df['Type'] == 'Pad'].copy() if 'Type' in filtered_df.columns else pd.DataFrame()
            
            if not pad_df.empty:
                # Prepare data for bar chart with Grid ID
                grid_id_col = col_names.get('grid_id', 'Grid ID')
                pad_data = pad_df.groupby('Location')[pts_col].mean().reset_index()
                # Get first Grid ID value for each location
                grid_ids_pad = pad_df.groupby('Location')[grid_id_col].first().reset_index()
                pad_data = pad_data.merge(grid_ids_pad, on='Location')
                pad_data.columns = ['Pad ID', 'Point Count', 'Grid ID']
                
                # Define colors based on traffic light logic
                def get_health_color_pad(points):
                    if pd.isna(points):
                        return chart_colors.get('traffic_light_gray', '#999999')
                    green_threshold = optical_thresholds.get('green_min', 75)
                    orange_threshold = optical_thresholds.get('orange_min', 40)
                    if points > green_threshold:
                        return chart_colors.get('traffic_light_green', '#2ecc71')
                    elif points >= orange_threshold:
                        return chart_colors.get('traffic_light_orange', '#f39c12')
                    else:
                        return chart_colors.get('traffic_light_red', '#e74c3c')
                
                pad_data['Color'] = pad_data['Point Count'].apply(get_health_color_pad)
                
                # Create bar chart with Grid ID labels
                fig_pad_health = go.Figure(data=[
                    go.Bar(
                        x=pad_data['Pad ID'],
                        y=pad_data['Point Count'],
                        marker=dict(color=pad_data['Color']),
                        text=[f"{pts:.1f}<br>({gid})" for pts, gid in zip(pad_data['Point Count'], pad_data['Grid ID'])],
                        textposition='outside',
                        customdata=pad_data['Grid ID'],
                        hovertemplate='<b>%{x}</b><br>Grid ID: %{customdata}<br>Points: %{y:.1f}<extra></extra>'
                    )
                ])
                
                fig_pad_health.update_layout(
                    title="Pad Health Status by Point Count",
                    xaxis_title="Pad ID",
                    yaxis_title="Point Count (Pts.)",
                    height=chart_heights.get('health_status', 550),
                    showlegend=False,
                    plot_bgcolor=chart_colors.get('chart_background', '#FFFFFF'),
                    hovermode='x unified',
                    xaxis=dict(showgrid=chart_colors.get('chart_gridlines_visible', False)),
                    yaxis=dict(showgrid=chart_colors.get('chart_gridlines_visible', False))
                )
                
                st.plotly_chart(fig_pad_health, use_container_width=True)
            else:
                st.warning("No Pad data found in the dataset.")
            
            # --- 2. Pad vs Via Separation Scatter Plot ---
            st.markdown("### 2️⃣ Pad vs. Via Edge Detection Verification")
            
            if 'Type' in filtered_df.columns:
                # Prepare data with Location/Grid ID
                grid_id_col = col_names.get('grid_id', 'Grid ID')
                scatter_data = []
                for item_type in ['Pad', 'Via']:
                    type_data = filtered_df[filtered_df['Type'] == item_type]
                    for _, row in type_data.iterrows():
                        if pd.notna(row[pts_col]):
                            scatter_data.append({
                                'Type': item_type,
                                'Point Count': row[pts_col],
                                'Location': row.get('Location', 'N/A'),
                                'Grid ID': row.get(grid_id_col, 'N/A')
                            })
                
                scatter_df = pd.DataFrame(scatter_data)
                
                if not scatter_df.empty:
                    # Create box plot with scatter overlay
                    fig_comparison = go.Figure()
                    
                    # Map categories to numeric positions for jitter
                    category_to_num = {'Pad': 0, 'Via': 1}
                    
                    for idx, item_type in enumerate(['Pad', 'Via']):
                        type_subset = scatter_df[scatter_df['Type'] == item_type]
                        type_points = type_subset['Point Count']
                        
                        if len(type_points) > 0:
                            # Add box plot
                            fig_comparison.add_trace(go.Box(
                                x=[item_type] * len(type_points),
                                y=type_points,
                                name=item_type,
                                boxmean='sd',
                                marker_opacity=0.3,
                                boxpoints=False
                            ))
                            
                            # Add scatter with jitter - use numeric x positions
                            jitter = np.random.normal(0, 0.04, size=len(type_points))
                            x_numeric = np.array([category_to_num[item_type]] * len(type_points))
                            x_jittered = x_numeric + jitter
                            
                            # Create hover text with Location and Grid ID
                            hover_text = [f"{loc}<br>Grid: {gid}" for loc, gid in zip(type_subset['Location'], type_subset['Grid ID'])]
                            
                            fig_comparison.add_trace(go.Scatter(
                                x=x_jittered,
                                y=type_points,
                                mode='markers+text',
                                name=f'{item_type} (Individual)',
                                marker=dict(size=chart_markers.get('pad_via_scatter_size', 16), opacity=0.7),
                                text=type_subset['Grid ID'],
                                textposition='middle center',
                                textfont=dict(size=chart_markers.get('pad_via_text_size', 10), color='white'),
                                showlegend=False,
                                hovertemplate='<b>%{customdata}</b><br>Points: %{y:.0f}<extra></extra>',
                                customdata=hover_text
                            ))
                    
                    fig_comparison.update_xaxes(tickvals=[0, 1], ticktext=['Pad', 'Via'], showgrid=chart_colors.get('chart_gridlines_visible', False))
                    fig_comparison.update_layout(
                        title="Edge Detection Verification: Pads vs Vias",
                        yaxis_title="Point Count (Pts.)",
                        height=chart_heights.get('comparison_plot', 400),
                        plot_bgcolor=chart_colors.get('chart_background', '#FFFFFF'),
                        hovermode='closest',
                        boxmode='group',
                        yaxis=dict(showgrid=chart_colors.get('chart_gridlines_visible', False))
                    )
                    
                    st.plotly_chart(fig_comparison, use_container_width=True)
            
            # --- 3. Detailed Analysis Tables for Vias and Pads ---
            st.markdown("### 3️⃣ Optical Edge Detection Details")
            
            # --- Via Details Table ---
            st.markdown("#### 🔴 Via Analysis (All Vias by Lowest Point Count)")
            
            via_table_df = filtered_df[filtered_df['Type'] == 'Via'].copy().dropna(subset=[pts_col])
            via_table_df = via_table_df.sort_values(by=pts_col, ascending=True)
            
            if not via_table_df.empty:
                # Prepare columns
                grid_id_col = col_names.get('grid_id', 'Grid ID')
                ptv_col = col_names.get('ptv_distance', 'PtV Distance')
                dx_col = col_names.get('x_distance', 'Shift (DX)')
                dy_col = col_names.get('y_distance', 'Shift (DY)')
                
                via_display = via_table_df[['Location', grid_id_col, pts_col, ptv_col, dx_col, dy_col]].copy()
                
                # Convert to microns (multiply by 1000)
                if ptv_col in via_display.columns:
                    via_display[f'{ptv_col} (µm)'] = (via_display[ptv_col] * 1000).round(3)
                if dx_col in via_display.columns:
                    via_display[f'{dx_col} (µm)'] = (via_display[dx_col] * 1000).round(3)
                if dy_col in via_display.columns:
                    via_display[f'{dy_col} (µm)'] = (via_display[dy_col] * 1000).round(3)
                
                # Drop original mm columns and keep only µm versions
                cols_to_keep = ['Location', grid_id_col, pts_col, f'{ptv_col} (µm)', f'{dx_col} (µm)', f'{dy_col} (µm)']
                cols_to_keep = [col for col in cols_to_keep if col in via_display.columns]
                via_display = via_display[cols_to_keep]
                via_display.columns = ['Location', 'Grid ID', 'Pts.', 'PtV (µm)', 'Shift DX (µm)', 'Shift DY (µm)']
                
                st.dataframe(via_display, use_container_width=True, hide_index=True)
            else:
                st.info("No Via data available")
            
            st.markdown("---")
            
            # --- Pad Details Table ---
            st.markdown("#### 🟠 Pad Analysis (All Pads by Lowest Point Count)")
            
            pad_table_df = filtered_df[filtered_df['Type'] == 'Pad'].copy().dropna(subset=[pts_col])
            pad_table_df = pad_table_df.sort_values(by=pts_col, ascending=True)
            
            if not pad_table_df.empty:
                # Prepare columns
                grid_id_col = col_names.get('grid_id', 'Grid ID')
                ptv_col = col_names.get('ptv_distance', 'PtV Distance')
                dx_col = col_names.get('x_distance', 'Shift (DX)')
                dy_col = col_names.get('y_distance', 'Shift (DY)')
                
                pad_display = pad_table_df[['Location', grid_id_col, pts_col, ptv_col, dx_col, dy_col]].copy()
                
                # Convert to microns (multiply by 1000)
                if ptv_col in pad_display.columns:
                    pad_display[f'{ptv_col} (µm)'] = (pad_display[ptv_col] * 1000).round(3)
                if dx_col in pad_display.columns:
                    pad_display[f'{dx_col} (µm)'] = (pad_display[dx_col] * 1000).round(3)
                if dy_col in pad_display.columns:
                    pad_display[f'{dy_col} (µm)'] = (pad_display[dy_col] * 1000).round(3)
                
                # Drop original mm columns and keep only µm versions
                cols_to_keep = ['Location', grid_id_col, pts_col, f'{ptv_col} (µm)', f'{dx_col} (µm)', f'{dy_col} (µm)']
                cols_to_keep = [col for col in cols_to_keep if col in pad_display.columns]
                pad_display = pad_display[cols_to_keep]
                pad_display.columns = ['Location', 'Grid ID', 'Pts.', 'PtV (µm)', 'Shift DX (µm)', 'Shift DY (µm)']
                
                st.dataframe(pad_display, use_container_width=True, hide_index=True)
            else:
                st.info("No Pad data available")

        # --- Process Stability View ---
        elif sub_view == "Process Stability":
            st.subheader("🔧 Process Stability: Diameter Consistency Analysis")
            st.write("Monitor tool wear and etching consistency through diameter variances. Two separate scales prevent Via precision from being masked by Pad dimensions.")
            
            # Load process stability settings - use session state if available
            process_stability = settings.get('PROCESS_STABILITY', {})
            outer_diam_col = col_names.get('outer_diameter', 'Outer Diameter')
            location_col = col_names.get('location', 'Location')
            
            # Get user-configured limits from session state, or use defaults
            via_nominal = st.session_state.get('via_nominal', process_stability.get('via_nominal_diameter', 0.020))
            via_ucl = st.session_state.get('via_ucl', process_stability.get('via_ucl', 0.0215))
            via_lcl = st.session_state.get('via_lcl', process_stability.get('via_lcl', 0.0185))
            
            pad_nominal = st.session_state.get('pad_nominal', process_stability.get('pad_nominal_diameter', 0.3125))
            pad_ucl = st.session_state.get('pad_ucl', process_stability.get('pad_ucl', 0.3157))
            pad_lcl = st.session_state.get('pad_lcl', process_stability.get('pad_lcl', 0.3093))
            
            # Create two columns for side-by-side charts
            col1, col2 = st.columns(2)
            
            # --- CHART 1: Via Drill Consistency ---
            with col1:
                st.markdown("### 🔴 Via Drill Consistency")
                st.caption("Mechanical/Laser Health - Tight control ±1.5µm")
                
                # Filter for Vias - check if 'Type' column exists and has 'Via' values
                if 'Type' in filtered_df.columns:
                    via_df = filtered_df[filtered_df['Type'] == 'Via'].copy()
                else:
                    # If no Type column, try to identify Vias by Location naming or show all data
                    via_df = filtered_df.copy()
                
                if not via_df.empty and outer_diam_col in via_df.columns:
                    # Prepare data
                    via_plot_df = via_df[[location_col, outer_diam_col]].dropna().copy()
                    
                    if not via_plot_df.empty:
                        via_plot_df = via_plot_df.sort_values(by=location_col)
                        
                        # Convert diameter to microns for plotting
                        via_plot_df['diameter_um'] = via_plot_df[outer_diam_col] * 1000
                        
                        # Calculate Y-axis range dynamically based on user-configured limits
                        # Add 2 µm margin around the control limits
                        via_nominal_um = via_nominal * 1000
                        via_ucl_um = via_ucl * 1000
                        via_lcl_um = via_lcl * 1000
                        margin = 2  # µm margin
                        
                        y_min = min(via_lcl_um, via_plot_df['diameter_um'].min()) - margin
                        y_max = max(via_ucl_um, via_plot_df['diameter_um'].max()) + margin
                        
                        # Determine colors: red if outside limits, green if within
                        via_plot_df['color'] = via_plot_df[outer_diam_col].apply(
                            lambda x: process_stability.get('outlier_color', '#d62728') 
                            if (x < via_lcl or x > via_ucl) 
                            else process_stability.get('normal_color', '#2ca02c')
                        )
                        
                        fig_via = go.Figure()
                        
                        # Add scatter plot (Y-axis in µm)
                        fig_via.add_trace(go.Scatter(
                            x=via_plot_df[location_col],
                            y=via_plot_df['diameter_um'],
                            mode='markers',
                            marker=dict(
                                size=process_stability.get('marker_size', 8),
                                color=via_plot_df['color'],
                                line=dict(width=1, color='rgba(0,0,0,0.3)')
                            ),
                            text=via_plot_df[location_col],
                            hovertemplate='<b>%{text}</b><br>Diameter: %{y:.1f} µm<extra></extra>',
                            showlegend=False
                        ))
                        
                        # Add nominal line (convert to µm scale)
                        fig_via.add_hline(
                            y=via_nominal * 1000,
                            line_dash="dash",
                            line_color=process_stability.get('nominal_color', '#1f77b4'),
                            line_width=process_stability.get('nominal_width', 2),
                            annotation_text=f"Nominal: {via_nominal*1000:.1f} µm",
                            annotation_position="right"
                        )
                        
                        # Add UCL line (convert to µm scale)
                        fig_via.add_hline(
                            y=via_ucl * 1000,
                            line_dash="dot",
                            line_color=process_stability.get('control_limit_color', '#ff7f0e'),
                            line_width=process_stability.get('control_limit_width', 2),
                            annotation_text=f"UCL: {via_ucl*1000:.1f} µm",
                            annotation_position="right"
                        )
                        
                        # Add LCL line (convert to µm scale)
                        fig_via.add_hline(
                            y=via_lcl * 1000,
                            line_dash="dot",
                            line_color=process_stability.get('control_limit_color', '#ff7f0e'),
                            line_width=process_stability.get('control_limit_width', 2),
                            annotation_text=f"LCL: {via_lcl*1000:.1f} µm",
                            annotation_position="right"
                        )
                        
                        fig_via.update_layout(
                            title="Via Outer Diameter Distribution",
                            xaxis_title="Target ID",
                            yaxis_title="Diameter (µm)",
                            height=500,
                            plot_bgcolor=chart_colors.get('chart_background', '#FFFFFF'),
                            hovermode='closest',
                            yaxis=dict(
                                range=[y_min, y_max],
                                showgrid=chart_colors.get('chart_gridlines_visible', False)
                            ),
                            xaxis=dict(
                                showgrid=chart_colors.get('chart_gridlines_visible', False)
                            )
                        )
                        
                        st.plotly_chart(fig_via, use_container_width=True)
                        
                        # Summary stats
                        out_of_spec = (via_plot_df[outer_diam_col] < via_lcl) | (via_plot_df[outer_diam_col] > via_ucl)
                        n_out = out_of_spec.sum()
                        pct_out = (n_out / len(via_plot_df) * 100) if len(via_plot_df) > 0 else 0
                        
                        st.metric("Out of Spec", f"{n_out} / {len(via_plot_df)} ({pct_out:.1f}%)", 
                                 delta="⚠️ Review needed" if n_out > 0 else "✓ All Pass", 
                                 delta_color="inverse" if n_out > 0 else "off")
                    else:
                        st.warning("No Via diameter data available after filtering NaN values")
                else:
                    st.warning("No Via diameter data available")
            
            # --- CHART 2: Pad Etch Consistency ---
            with col2:
                st.markdown("### 🟠 Pad Etch Consistency")
                st.caption("Chemistry Health - Tight control ±3.2µm")
                
                # Filter for Pads - check if 'Type' column exists
                if 'Type' in filtered_df.columns:
                    pad_df = filtered_df[filtered_df['Type'] == 'Pad'].copy()
                else:
                    pad_df = filtered_df.copy()
                
                if not pad_df.empty and outer_diam_col in pad_df.columns:
                    # Prepare data
                    pad_plot_df = pad_df[[location_col, outer_diam_col]].dropna().copy()
                    
                    if not pad_plot_df.empty:
                        pad_plot_df = pad_plot_df.sort_values(by=location_col)
                        
                        # Convert diameter to microns for plotting
                        pad_plot_df['diameter_um'] = pad_plot_df[outer_diam_col] * 1000
                        
                        # Calculate Y-axis range dynamically based on user-configured limits
                        # Add 2 µm margin around the control limits
                        pad_nominal_um = pad_nominal * 1000
                        pad_ucl_um = pad_ucl * 1000
                        pad_lcl_um = pad_lcl * 1000
                        margin = 2  # µm margin
                        
                        y_min = min(pad_lcl_um, pad_plot_df['diameter_um'].min()) - margin
                        y_max = max(pad_ucl_um, pad_plot_df['diameter_um'].max()) + margin
                        
                        # Determine colors: red if outside limits, green if within
                        pad_plot_df['color'] = pad_plot_df[outer_diam_col].apply(
                            lambda x: process_stability.get('outlier_color', '#d62728') 
                            if (x < pad_lcl or x > pad_ucl) 
                            else process_stability.get('normal_color', '#2ca02c')
                        )
                        
                        fig_pad = go.Figure()
                        
                        # Add scatter plot (Y-axis in µm)
                        fig_pad.add_trace(go.Scatter(
                            x=pad_plot_df[location_col],
                            y=pad_plot_df['diameter_um'],
                            mode='markers',
                            marker=dict(
                                size=process_stability.get('marker_size', 8),
                                color=pad_plot_df['color'],
                                line=dict(width=1, color='rgba(0,0,0,0.3)')
                            ),
                            text=pad_plot_df[location_col],
                            hovertemplate='<b>%{text}</b><br>Diameter: %{y:.1f} µm<extra></extra>',
                            showlegend=False
                        ))
                        
                        # Add nominal line (convert to µm scale)
                        fig_pad.add_hline(
                            y=pad_nominal * 1000,
                            line_dash="dash",
                            line_color=process_stability.get('nominal_color', '#1f77b4'),
                            line_width=process_stability.get('nominal_width', 2),
                            annotation_text=f"Nominal: {pad_nominal*1000:.1f} µm",
                            annotation_position="right"
                        )
                        
                        # Add UCL line (convert to µm scale)
                        fig_pad.add_hline(
                            y=pad_ucl * 1000,
                            line_dash="dot",
                            line_color=process_stability.get('control_limit_color', '#ff7f0e'),
                            line_width=process_stability.get('control_limit_width', 2),
                            annotation_text=f"UCL: {pad_ucl*1000:.1f} µm",
                            annotation_position="right"
                        )
                        
                        # Add LCL line (convert to µm scale)
                        fig_pad.add_hline(
                            y=pad_lcl * 1000,
                            line_dash="dot",
                            line_color=process_stability.get('control_limit_color', '#ff7f0e'),
                            line_width=process_stability.get('control_limit_width', 2),
                            annotation_text=f"LCL: {pad_lcl*1000:.1f} µm",
                            annotation_position="right"
                        )
                        
                        fig_pad.update_layout(
                            title="Pad Outer Diameter Distribution",
                            xaxis_title="Target ID",
                            yaxis_title="Diameter (µm)",
                            height=500,
                            plot_bgcolor=chart_colors.get('chart_background', '#FFFFFF'),
                            hovermode='closest',
                            yaxis=dict(
                                range=[y_min, y_max],
                                showgrid=chart_colors.get('chart_gridlines_visible', False)
                            ),
                            xaxis=dict(
                                showgrid=chart_colors.get('chart_gridlines_visible', False)
                            )
                        )
                        
                        st.plotly_chart(fig_pad, use_container_width=True)
                        
                        # Summary stats
                        out_of_spec = (pad_plot_df[outer_diam_col] < pad_lcl) | (pad_plot_df[outer_diam_col] > pad_ucl)
                        n_out = out_of_spec.sum()
                        pct_out = (n_out / len(pad_plot_df) * 100) if len(pad_plot_df) > 0 else 0
                        
                        st.metric("Out of Spec", f"{n_out} / {len(pad_plot_df)} ({pct_out:.1f}%)", 
                                 delta="⚠️ Review needed" if n_out > 0 else "✓ All Pass", 
                                 delta_color="inverse" if n_out > 0 else "off")
                    else:
                        st.warning("No Pad diameter data available after filtering NaN values")
                else:
                    st.warning("No Pad diameter data available")

    elif main_view == "Via to Pad":
        st.info("Via to Pad analytics not yet implemented.")

    elif main_view == "Alignment":
        st.info("Alignment analytics not yet implemented.")

if __name__ == "__main__":
    main()
