"""
Quality Control View - Statistical summary and annular ring analysis
Extracted from app.py for modularity and reusability.
"""

import streamlit as st
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill

from src.views.base import BaseView


class QualityControlView(BaseView):
    """
    QualityControlView: Displays statistical summary and detailed quality metrics.
    
    Features:
    - Statistical summary: min, max, mean, std of annular ring
    - Failure counts and percentages
    - Excel export with conditional highlighting
    - Interactive dataframe with custom formatting
    """
    
    def render(self, filtered_df: pd.DataFrame, **kwargs) -> None:
        """
        Render the Quality Control analysis view.
        
        Args:
            filtered_df: Input DataFrame with measurement data
            **kwargs: Additional context including 'col_names' and 'tolerances'
        """
        col_names = kwargs.get('col_names', {})
        tolerances = kwargs.get('tolerances', {})
        
        annular_col = col_names.get('annular_ring', 'Annular Ring')
        threshold = tolerances.get('annular_ring_min', 0.0)
        grid_id_col = col_names.get('grid_id', 'Grid ID')
        
        # Prepare display dataframe
        display_df = filtered_df.copy()
        
        # Drop Panel and Side columns (not needed in display)
        cols_to_drop = [col for col in ['Panel', 'Side'] if col in display_df.columns]
        display_df = display_df.drop(columns=cols_to_drop)
        
        # Convert all measurements from mm to microns (multiply by 1000)
        measurement_cols = [
            col_names.get('outer_diameter', 'Outer Diameter'),
            col_names.get('inner_diameter', 'Inner Diameter'),
            col_names.get('ptv_distance', 'PtV Distance'),
            col_names.get('total_shift', 'SC'),
            col_names.get('x_distance', 'Shift (DX)'),
            col_names.get('y_distance', 'Shift (DY)'),
            col_names.get('annular_ring', 'Annular Ring')
        ]
        for col in measurement_cols:
            if col in display_df.columns:
                display_df[col] = (display_df[col] * 1000).round(3)
        
        # Sort by Grid ID ascending
        if grid_id_col in display_df.columns:
            display_df = display_df.sort_values(by=grid_id_col, ascending=True)
        
        # --- Display Statistical Summary ---
        st.subheader("📊 Statistical Summary")
        summary_cols = st.columns(5)
        
        # Calculate stats for annular ring
        if annular_col in display_df.columns:
            ar_min = display_df[annular_col].min()
            ar_max = display_df[annular_col].max()
            ar_mean = display_df[annular_col].mean()
            ar_std = display_df[annular_col].std()
            fail_count = (display_df[annular_col] < threshold).sum()
            fail_pct = (fail_count / len(display_df) * 100) if len(display_df) > 0 else 0
            
            with summary_cols[0]:
                st.metric(
                    "Min Annular Ring",
                    f"{ar_min:.3f} µm",
                    delta=f"{ar_min - threshold:.3f}" if ar_min < threshold else None,
                    delta_color="inverse"
                )
            with summary_cols[1]:
                st.metric("Max Annular Ring", f"{ar_max:.3f} µm")
            with summary_cols[2]:
                st.metric("Mean ± StdDev", f"{ar_mean:.3f} ± {ar_std:.3f} µm")
            with summary_cols[3]:
                worst_idx = display_df[annular_col].idxmin()
                worst_grid = display_df.loc[worst_idx, grid_id_col] if grid_id_col in display_df.columns else 'N/A'
                st.metric("Worst Point", f"Grid {worst_grid}", f"{ar_min:.3f} µm")
            with summary_cols[4]:
                st.metric(
                    "Failures",
                    f"{fail_count} ({fail_pct:.1f}%)",
                    delta=f"{fail_count} points" if fail_count > 0 else "✓ All Pass",
                    delta_color="inverse" if fail_count > 0 else "off"
                )
        
        st.markdown("---")
        
        # --- Export to Excel button ---
        col_left, col_right = st.columns([3, 1])
        with col_left:
            st.write(f"Highlighting rows where {annular_col} < {threshold}")
        with col_right:
            # Create Excel file in memory
            buffer = BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                # Write data with highlighting
                display_df_export = display_df.copy()
                display_df_export.to_excel(writer, sheet_name='QVM Data', index=False)
                
                # Get workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['QVM Data']
                
                # Apply conditional formatting for failed points
                red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                
                if annular_col in display_df_export.columns:
                    ar_col_idx = list(display_df_export.columns).index(annular_col) + 1  # Excel is 1-indexed
                    for row_idx, value in enumerate(display_df_export[annular_col], start=2):  # start=2 to skip header
                        if value < threshold:
                            for col_idx in range(1, len(display_df_export.columns) + 1):
                                worksheet.cell(row=row_idx, column=col_idx).fill = red_fill
            
            buffer.seek(0)
            st.download_button(
                label="📥 Export to Excel",
                data=buffer,
                file_name=f"QVM_Quality_Report_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )
        
        # Add clarification about Grid ID
        st.info(
            "📍 **Grid ID**: Position on the PCB quadrant (11-14 = Upper Left, "
            "21-24 = Lower Left, 31-34 = Lower Right, 41-44 = Upper Right). "
            "First digit = quadrant, second digit = point (1-4). "
            "All measurements in **microns (µm)**."
        )
        
        # --- Display Interactive Dataframe ---
        styled_df = display_df.style.apply(
            self._highlight_annular_ring,
            col_name=annular_col,
            threshold=threshold,
            axis=1
        )
        
        # Apply custom formatting for 3 decimals
        format_dict = {}
        for col in display_df.columns:
            if display_df[col].dtype in ['float64', 'float32', 'int64', 'int32']:
                format_dict[col] = self.format_to_3_decimals
        
        if format_dict:
            styled_df = styled_df.format(format_dict)
        
        st.dataframe(styled_df, use_container_width=True)
    
    @staticmethod
    def _highlight_annular_ring(row, col_name, threshold):
        """
        Styling function to highlight rows with low annular ring.
        
        Args:
            row: DataFrame row
            col_name: Name of annular ring column
            threshold: Failure threshold
            
        Returns:
            List of CSS styles for each cell in the row
        """
        if col_name not in row.index:
            return [''] * len(row)
        
        value = row[col_name]
        if pd.isna(value):
            return [''] * len(row)
        
        # Highlight row in light red if value is below threshold
        if value < threshold:
            return ['background-color: #ffcccc'] * len(row)
        return [''] * len(row)
